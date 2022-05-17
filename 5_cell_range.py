from random import randint
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string


wb = Workbook()
ws = wb.active

# 한줄씩 데이터 넣기
ws.append(["번호","영어","수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# 지정컬럼의 데이터만 추출
col_B = ws["B"]
for cell in col_B:
    print(cell.value, end=" ")
print()
print()

# 지정 컬럼 범위 데이터 추출
col_range = ws["B:C"]
for col in col_range:
    for cell in col:
        print(cell.value, end=" ")
    print()
print()
print()

# 지정 로우만 추출
row_title = ws[1]
for row in row_title:
    print(row.value, end=" ")
print()
print()

# 지정 로우 범위 데이터 추출
rows_data = ws[2:6]     # 2 ~ 6 번째 줄까지 추출
for col in rows_data:
    for row in col:
        print(row.value, end=" ")
    print()
print()
print()


rows_data = ws[2:ws.max_row]     # 2 ~ 마지막 번째 줄까지 추출
for col in rows_data:
    for row in col:
        print(row.value, end="/ ")
        print(row.coordinate, end="/ ")      # 셀 위치 추출
        
        xy = coordinate_from_string(row.coordinate)     # 셀위치를 튜플로 반환
        print(xy, end=" ")

    print()
print()
print()


print(tuple(ws.rows))               # 로우순서로(가로) 읽어 튜플로 반환
print(tuple(ws.columns))            # 컬럼순서로(세로) 읽어 튜플로 반환
for row in tuple(ws.rows):
    print(row[0].value, end=" ")
    print(row[1].value, end=" ")
    print(row[2].value)
    print()


# 원하는 범위를 설정해서 추출 (ws.iter_rows) : 로우기준
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value, row[1].value)
print()
print()

# 원하는 범위를 설정해서 추출 (ws.iter_cols) : 컬럼기준
for col in ws.iter_cols(min_col=2, max_col=3, min_row=1, max_row=5):
    print(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value)

wb.save("sample4.xlsx")
wb.close