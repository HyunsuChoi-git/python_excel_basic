from sqlite3 import Row
from openpyxl import load_workbook      # 파일 불러오기


wb = load_workbook("sample3.xlsx")       # sample.xlsx 파일에서 wb을 불러옴
ws = wb.active                          # 활성화된 시트 가져옴

#cell 데이터 불러오기

for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()


#cell 정보를 모를 때 ,
for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
