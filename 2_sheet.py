from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()          # 새로운 Sheet 기본 이름으로 생성 (Sheet)
ws.title = "MySheet"            # Sheet명 변경
ws.sheet_properties.tabColor = "ff66ff"   # RGB형태로 값을 넣어주면 Sheet명의 배경색이 변경된다.

ws1 = wb.create_sheet("YourSheet")      # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)    # Sheet 순서를 지정하여 생성


new_ws = wb["NewSheet"]             # dic형태로 sheet에 접근 가능

print(wb.sheetnames)

new_ws["A1"] = "Test"           # 해당 셀에 데이터 입력
target = wb.copy_worksheet(new_ws)        # sheet 복사
target.title = "copied sheet"   # 복사한 시트 sheet명 변경

wb.save("sample2.xlsx")