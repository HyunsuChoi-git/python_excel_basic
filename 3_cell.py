from random import randint
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "HeraSheet"

ws["A1"] = 1            # A1셀에 1 입력
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"].value)       # .value  -> 셀의 값을 추출
print(ws["A2"].value)       # .value  -> 셀의 값을 추출
print(ws["A10"].value)       # 값이 없을 땐 None

print(ws.cell(row=1, column=1).value)     # == ws["A1"].value
c = ws.cell(row=1, column=3, value=10)      # C1에 10 입력
print(c.value)

# 임의의 숫자 엑셀에 입력하기
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0, 100))



wb.save("sample3.xlsx")
wb.close